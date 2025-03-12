import datetime
import io
import os
import random

from django.http import HttpResponse, HttpRequest
from django.shortcuts import render, redirect
import openpyxl
from django.utils import timezone
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.models import Course, Enrollment, EnrollmentTask, Screenshot, Task, TaskType


# also make a report with percentage, merge by school if same columns

def reports(request: HttpRequest) -> HttpResponse:
    if request.user.is_anonymous or not request.user.is_staff:
        return redirect("/")

    return render(request, "reports.html", {
        "courses": Course.objects.filter(teacher=request.user).filter(deleted=False),
        "task_types": TaskType.objects.all().order_by('name'),
        "page": "reports",
    })


def filter_letters(in_str: str) -> str:
    return "".join([letter for letter in in_str if letter.isalnum() or letter == " "])


def course_to_journal(
        course: Course,
        show_each_task: bool = True,
        show_notes: bool = True,
        metric_percent: bool = False,
        group_by_department: bool = False,
        **kwargs,
) -> tuple[list[str], list[list[str]], list[Task], int]:
    """
    input: Course
    output: header, journal_lines, tasks
    """
    shift = 4
    tasks = list(course.task_set.order_by('order').all())
    now = timezone.now()
    
    # Filter by task type if specified
    task_type_id = kwargs.get("task_type")
    if task_type_id:
        tasks = [task for task in tasks if task.task_type_id == int(task_type_id)]
    
    # Convert date strings to datetime objects if they exist
    date_from = None
    date_to = None
    if kwargs.get("date_from"):
        date_from = timezone.datetime.strptime(kwargs["date_from"], "%Y-%m-%d")
        date_from = timezone.make_aware(date_from)
    if kwargs.get("date_to"):
        date_to = timezone.datetime.strptime(kwargs["date_to"], "%Y-%m-%d")
        date_to = timezone.make_aware(date_to)

    if date_from is not None and date_to is None:
        date_to = now
    if date_from is None and date_to is not None:                
        date_from = timezone.datetime(2024, 1, 1)
        date_from = timezone.make_aware(date_from)

    # First filter tasks by availability
    tasks = [
        task
        for task in tasks
        if task.available_from is None or task.available_from <= now
    ]

    # This we should use if we want to really see if there are screenshots
    # Then filter tasks by date range if specified
    # if date_from is not None and date_to is not None:
    #     for task in tasks:
    #         task.should_be_removed = False
    #         enrollment_tasks = EnrollmentTask.objects.filter(task=task, enrollment__course=course)
    #         has_screenshots = False
    #         for enrollment_task in enrollment_tasks:
    #             screenshots = Screenshot.objects.filter(
    #                 enrollment_task=enrollment_task,
    #                 created_at__gte=date_from,
    #                 created_at__lte=date_to
    #             ).exists()
    #             if screenshots:
    #                 has_screenshots = True
    #                 break
    #         if not has_screenshots:
    #             task.should_be_removed = True

    #     tasks = [task for task in tasks if not task.should_be_removed]

    if date_from is not None:
        tasks = [task for task in tasks if task.available_from is None or task.available_from >= date_from]

    if date_to is not None:
        tasks = [task for task in tasks if task.available_from is None or task.available_from <= date_to]

    enrollments = Enrollment.objects.filter(course=course, deleted=False).order_by(
        'student__last_name',
        'student__first_name',
    ).prefetch_related("student")
    enrollment_tasks = EnrollmentTask.objects.filter(enrollment__course=course).prefetch_related("enrollment")
    journal_lines: list[list[str]] = list()
    header: list[str] = list()
    if group_by_department:
        header.append("Course")
        shift += 1
    header.append("Student")
    header.append("Email")
    if metric_percent:
        header.append("Percent complete")
    else:
        header.append("Total minutes")
    if show_each_task:
        for task in tasks:
            header.append(task.name)
            if show_notes:
                header.append("")

    if len(tasks) == 0:
        return header, [], [], shift

    tasks_total_time = sum([
        task1.minimum_minutes
        for task1 in tasks
    ])
    if tasks_total_time == 0:
        tasks_total_time = 1

    task_ids = [task.id for task in tasks]

    for enrollment in enrollments:
        line = list()
        if group_by_department:
            line.append(f"{course.name}")
        line.append(f"{enrollment.student.last_name} {enrollment.student.first_name}")
        line.append(enrollment.student.email)
        if metric_percent:
            perc = enrollment.total_minutes_func(tasks=tasks) / tasks_total_time
            perc = int(perc*100)/100
            line.append(perc)
        else:
            line.append(enrollment.total_minutes_func(tasks=tasks))
        if show_each_task:
            for task in tasks:
                time = ""
                note = ""
                for enrollment_task in enrollment_tasks:
                    if enrollment_task.enrollment == enrollment and enrollment_task.task == task:
                        if metric_percent:
                            if task.minimum_minutes > 0:
                                perc = enrollment_task.minutes / task.minimum_minutes
                                perc = int(perc*100)/100
                                time = perc
                            else:
                                time = "divBy0"
                        else:
                            # hours = str(enrollment_task.minutes // 60).zfill(2)
                            # minutes = str(enrollment_task.minutes % 60).zfill(2)
                            time = enrollment_task.minutes
                        note = enrollment_task.note
                line.append(time)
                if show_notes:
                    line.append(note)
        journal_lines.append(line)
    return header, journal_lines, tasks, shift


def export_all(request):
    # Validate user access first
    if not request.user.is_authenticated or not request.user.is_staff:
        return HttpResponse("Unauthorized", status=401)

    # Get and validate parameters
    group_by_department = request.GET.get("groupByDepartment", "false") == "true"
    show_each_task = request.GET.get("showEachTask", "false") == "true"
    show_notes = request.GET.get("showNotes", "false") == "true"
    metric_percent = request.GET.get("metric", "time") == "percentage"
    
    # Build kwargs for filtering
    kwargs = {}
    
    # Date handling
    date_from = request.GET.get("dateFrom", "").strip()
    date_to = request.GET.get("dateTo", "").strip()
    if date_from:
        kwargs["date_from"] = date_from
    if date_to:
        kwargs["date_to"] = date_to

    # Task type handling
    task_type_id = request.GET.get("taskType", "").strip()
    if task_type_id and task_type_id.isdigit():
        kwargs["task_type"] = task_type_id

    # Create workbook and setup initial sheet
    workbook = openpyxl.Workbook()
    sheet0 = workbook.active
    sheet0.title = "Summary"
    sheet0.append(["Export date", datetime.datetime.now().strftime("%d.%m.%Y")])
    sheet0.append(["Courses:"])

    # Track totals
    total_courses = 0
    total_minutes = 0
    department_sheets = {}

    try:
        # Process each course
        for course in Course.objects.filter(deleted=False):
            if not course.permissions(request.user).read:
                continue

            department = course.department.name if course.department else "No Department"
            
            header, journal_lines, tasks, shift = course_to_journal(
                course,
                show_each_task=show_each_task,
                show_notes=show_notes,
                metric_percent=metric_percent,
                group_by_department=group_by_department,
                **kwargs,
            )

            # Skip if no data
            if not journal_lines:
                continue

            # Update totals
            course_total_minutes = course.total_minutes(tasks=tasks)
            total_courses += 1
            total_minutes += course_total_minutes
            
            # Add to summary sheet
            hours = str(course_total_minutes // 60)
            minutes = str(course_total_minutes % 60).zfill(2)
            sheet0.append(["", department, course.name, f"{hours} h {minutes} min"])

            # Create or get appropriate sheet
            if group_by_department:
                if department in department_sheets:
                    sheet = department_sheets[department]
                else:
                    sheet = workbook.create_sheet(title=filter_letters(department)[:30])
                    department_sheets[department] = sheet
                    sheet.append(header)
            else:
                sheet = workbook.create_sheet(title=filter_letters(course.name)[:30])
                sheet.append(header)

            # Add data and format sheet
            for line in journal_lines:
                sheet.append(line)

            # Apply formatting
            for j in range(sheet.max_row - 1):
                if metric_percent:
                    sheet.cell(j + 2, shift - 1).style = 'Percent'

            # Format task columns
            for i in range(len(tasks)):
                if show_notes:
                    sheet.merge_cells(
                        start_row=1,
                        start_column=shift + i * 2,
                        end_row=1,
                        end_column=shift + i * 2 + 1
                    )
                    current_cell = sheet.cell(1, shift + i * 2)
                    current_cell.alignment = Alignment(horizontal='center')
                    current_cell.font = Font(bold=True)
                    sheet.column_dimensions[get_column_letter(shift + i * 2)].width = 10
                    sheet.column_dimensions[get_column_letter(shift + i * 2 + 1)].width = 20
                    
                    if metric_percent:
                        for j in range(len(journal_lines)):
                            sheet.cell(j + 2, shift + i * 2).style = 'Percent'
                else:
                    current_cell = sheet.cell(1, shift + i)
                    current_cell.alignment = Alignment(horizontal='center')
                    current_cell.font = Font(bold=True)
                    sheet.column_dimensions[get_column_letter(shift + i)].width = 20
                    
                    if metric_percent:
                        for j in range(sheet.max_row - 1):
                            sheet.cell(j + 2, shift + i).style = 'Percent'

            # Format header columns
            for i in range(shift):
                sheet.column_dimensions[get_column_letter(i + 1)].width = 20
                current_cell = sheet.cell(1, i + 1)
                current_cell.font = Font(bold=True)

        # Add totals to summary sheet
        sheet0.append(["Total:", f"{total_courses} courses"])
        hours = str(total_minutes // 60)
        minutes = str(total_minutes % 60).zfill(2)
        sheet0.append(["Total time:", f"{hours} h {minutes} min"])

        # Format summary sheet
        for i in range(1, 5):
            sheet0.column_dimensions[get_column_letter(i)].width = 20

        # Generate response
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="journal.xlsx"'
        return response

    except Exception as e:
        return HttpResponse(f"Error generating report: {str(e)}", status=500)


def journal(request, course_id):
    course = Course.objects.get(id=course_id)
    if not course.permissions(user=request.user).read:
        return redirect("/")

    header, journal_lines, tasks, shift = course_to_journal(course)

    if request.GET.get("to_excel") == "1":
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(header)
        for line in journal_lines:
            sheet.append(line)
        for i in range(len(tasks)):
            sheet.merge_cells(start_row=1, start_column=shift+i*2, end_row=1, end_column=shift+i*2+1)
            currentCell = sheet.cell(1,shift+i*2)
            currentCell.alignment = Alignment(horizontal='center')
            currentCell.font = Font(bold=True)
            sheet.column_dimensions[get_column_letter(shift+i*2)].width = 10
            sheet.column_dimensions[get_column_letter(shift+i*2+1)].width = 20
        sheet.column_dimensions[get_column_letter(1)].width = 20
        sheet.column_dimensions[get_column_letter(2)].width = 20
        sheet.column_dimensions[get_column_letter(3)].width = 20
        currentCell = sheet.cell(1, 1)
        currentCell.font = Font(bold=True)
        currentCell = sheet.cell(1, 2)
        currentCell.font = Font(bold=True)
        currentCell = sheet.cell(1, 3)
        currentCell.font = Font(bold=True)
        buffer = io.BytesIO()
        workbook.save(buffer)

        buffer.seek(0)
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="journal.xlsx"'
        return response

    return render(
        request,
        "journal.html",
        {
            'course': course,
            'courses': Course.objects.filter(teacher=request.user).filter(deleted=False),
            'other_courses': Course.objects.filter(teacher=request.user).exclude(id=course_id).filter(deleted=False),
            'tasks': tasks,
            'page': 'journal',
            'header': header,
            'journal_lines': journal_lines,
        }
    )
